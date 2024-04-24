import os
import pandas as pd

import pytz
import datetime

import openpyxl
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.worksheet.table import TableStyleInfo

# Create a table style
table_style = TableStyleInfo(
    name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False
)


def initialize_workbook():
    workbook = load_workbook(os.path.join("report", "template", "workbook.xlsx"))
    workbook.save(os.path.join("report", "InfoMagnus - Migration Workbook.xlsx"))


def get_workbook(workbook_path):
    workbook = load_workbook(workbook_path)
    workbook.filename = workbook_path
    print(f"** Found workbook at: {workbook_path}")

    return workbook


def get_included_orgs(org_type, workbook_path):
    # Load the Org Mappings
    wb = load_workbook(workbook_path, data_only=True)

    ws = wb["Mapping - Org"]

    data = list(ws.values)

    # Set the first row as the header
    df = pd.DataFrame(data[1:], columns=data[0])

    # Filter out excluded orgs
    orgs = df[df["exclude"] == False][org_type].tolist()

    if orgs == ():
        raise ValueError("No source orgs found in 'Mapping - Org'")

    return orgs


def get_included_orgs_by_wave(org_type, wave, workbook_path):
    # Load the Org Mappings
    wb = load_workbook(workbook_path, data_only=True)

    ws = wb["Mapping - Org"]

    data = list(ws.values)

    # Set the first row as the header
    df = pd.DataFrame(data[1:], columns=data[0])

    # Get orgs for wave, filter out excluded orgs
    orgs = df[(df["exclude"] == False) & (df["wave"] == wave)][org_type].tolist()

    # If orgs is empty
    if orgs == ():
        raise ValueError("No source orgs found in 'Mapping - Org'")

    return orgs


def get_included_orgs_by_wave_df(org_type, wave, workbook_path):
    # Load the Org Mappings
    wb = load_workbook(workbook_path, data_only=True)

    ws = wb["Mapping - Org"]

    data = list(ws.values)

    # Set the first row as the header
    df = pd.DataFrame(data[1:], columns=data[0])

    # Get orgs for wave, filter out excluded orgs
    orgs = df[(df["exclude"] == False) & (df["wave"] == wave)]

    # If orgs is empty
    if orgs.empty:
        raise ValueError("No source orgs found in 'Mapping - Org'")

    return orgs


def add_sheet(workbook, sheet_name, desired_index=0, tab_color="FF0000"):
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.create_sheet(sheet_name, index=desired_index)
        worksheet.sheet_properties.tabColor = tab_color

    return worksheet


def autosize_columns(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column = column[0].column_letter
        for cell in worksheet[column]:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        adjusted_width = min(adjusted_width, 60)
        worksheet.column_dimensions[column].width = adjusted_width


def write_table(worksheet, df, table_name, heading=""):
    # If pushedAt exists in the df
    if "pushedAt" in df.columns:
        df["pushedAt"] = df["pushedAt"].dt.tz_localize(None)
    if "updatedAt" in df.columns:
        df["updatedAt"] = df["updatedAt"].dt.tz_localize(None)

    # Add a header with the table name
    if heading != "":
        worksheet.append([heading])
        worksheet[f"A{worksheet.max_row}"].font = Font(bold=True, size=12)

    # If there are no rows, add a single row with "No data"
    if df.empty:
        worksheet.append(["None"])
        worksheet[f"A{worksheet.max_row}"].font = Font(italic=True)
    else:
        # Add the table
        from openpyxl.worksheet.table import Table, TableStyleInfo

        num_rows, num_cols = df.shape
        if heading != "":
            start_col, start_row = "A", worksheet.max_row + 1
        else:
            start_col, start_row = "A", worksheet.max_row
        end_col = openpyxl.utils.get_column_letter(num_cols)
        end_row = start_row + num_rows

        table = Table(
            displayName=table_name,
            ref=f"{start_col}{start_row}:{end_col}{end_row}",
            tableStyleInfo=table_style,
        )
        worksheet.add_table(table)

        # Add the data
        worksheet.append(df.columns.to_list())
        for row in df.itertuples(index=False, name=None):
            worksheet.append(row)

        # Group added rows
        worksheet.row_dimensions.group(start_row + 1, end_row, hidden=False)
        autosize_columns(worksheet)

    # Move to the next empty row
    worksheet.append([])


def add_inventory_worksheet(workbook, sheet_name, stats):
    """ """
    # Delete the worksheet if it already exists
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    desired_index = workbook.sheetnames.index("Cover") + 1
    worksheet = add_sheet(workbook, sheet_name, desired_index, "002060")

    # Clear the contents of the worksheet
    worksheet.delete_rows(1, worksheet.max_row)

    write_table(worksheet, stats, "Inventory")


def add_worksheet(workbook, sheet_name, stats):
    # Delete the worksheet if it already exists
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    desired_index = workbook.sheetnames.index("Cover") + 1
    worksheet = add_sheet(workbook, sheet_name, desired_index, "002060")

    # Clear the contents of the worksheet
    worksheet.delete_rows(1, worksheet.max_row)

    # Create table name by removing spaces and adding underscores
    table_name = sheet_name.replace(" ", "_")

    write_table(worksheet, stats, table_name)


def write_mappings_file(df, cols):
    df = df[cols]

    # Rename the "name" columns to "source name" and "target name"
    cols = list(df.columns)

    # Rename the "name" columns to "source name" and "target name"
    if cols[0] == "name" and cols[1] == "name":
        cols[0] = "source_name"
        cols[1] = "target_name"
        cols[2] = "dry_run_target_name"
    else:
        raise ValueError('Columns 0 and 1 must be "name"')

    df.columns = cols
    return df


def add_org_mapping(workbook, sheet_name, stats):
    """ """
    # Delete the worksheet if it already exists
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    desired_index = workbook.sheetnames.index("Cover") + 2
    worksheet = add_sheet(workbook, sheet_name, desired_index, "FF0000")

    # Clear the contents of the worksheet
    worksheet.delete_rows(1, worksheet.max_row)

    # TODO: Clean up this mess...
    stats = stats.rename(columns={"name": "_name"})
    stats = stats.rename(columns={"owner.login": "name"})
    stats["wave"] = 0
    stats["order"] = 0
    stats["exclude"] = False
    stats["exclude_reason"] = ""

    # Remove dupes from stats
    stats = stats.drop_duplicates(subset=["name"])

    stats = write_mappings_file(
        stats, ["name", "name", "name", "wave", "order", "exclude", "exclude_reason"]
    )

    # Create org mapping table
    write_table(worksheet, stats, "Mapping_Org")

    workbook.save(workbook.filename)


def add_pre_migration_report(workbook, sheet_name, stats):
    """ """
    # Delete the worksheet if it already exists
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]

    desired_index = workbook.sheetnames.index("Cover") + 1
    worksheet = add_sheet(workbook, sheet_name, desired_index, "215C98")

    # Clear the contents of the worksheet
    worksheet.delete_rows(1, worksheet.max_row)

    # Create large repos table
    df = stats[stats["diskUsage"] > 1000].sort_values(by="diskUsage")
    write_table(worksheet, df, "Large_Repos", "Large Repos")

    # Create large PRs table
    df = stats[stats["pullRequests.totalCount"] > 1000].sort_values(
        by="pullRequests.totalCount"
    )
    write_table(worksheet, df, "Large_PRs", "Large PRs")

    # Create webhooks table
    df = stats[stats["webhooks.totalCount"] > 0].sort_values(by="webhooks.totalCount")
    write_table(worksheet, df, "Webhooks_Repos", "Repos with webhooks")

    # Create actions table
    df = stats[stats["lastWorkflowRun"] != None].sort_values("lastWorkflowRun")
    write_table(worksheet, df, "Actions_Repos", "Repos with actions")

    # Create stale repos table
    stats["pushedAt"] = stats["pushedAt"].dt.tz_localize("UTC")
    df = stats[
        stats["pushedAt"]
        < datetime.datetime.now(pytz.UTC) - datetime.timedelta(days=60)
    ].sort_values("pushedAt")
    write_table(worksheet, df, "Stale_Repos", "Stale Repos")

    # Create archived repos table
    df = stats[stats["isArchived"] == True].sort_values("isArchived")
    write_table(worksheet, df, "Archived_Repos", "Archived Repos")

    # Create locked repos table
    df = stats[stats["isLocked"] == True].sort_values("isLocked")
    write_table(worksheet, df, "Locked_Repos", "Locked Repos")

    # Create repos with packages table
    df = stats[stats["packages.totalCount"] > 0].sort_values("packages.totalCount")
    write_table(worksheet, df, "Has_Packages", "Repos with packages")

    # def identify_git_lfs():
    # # TODO: Need to figure out how to implement this

    workbook.save(workbook.filename)
